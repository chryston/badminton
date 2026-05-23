#!/usr/bin/env python3
"""
Import historical data from Badminton.xlsx into Supabase.

Usage:
    cd /path/to/repo
    SUPABASE_URL=... SUPABASE_SERVICE_ROLE_KEY=... python scripts/import_excel.py

Reads from: Badminton.xlsx (located two directories above this script)
Inserts into: venues, players, shuttle_batches tables

Safe to re-run: existing rows (matched by name / batch_name+owner_label) are skipped.
Prints a summary of what was inserted vs skipped.

NOTE: Historical sessions from the "Profit & Loss" sheet are NOT imported.
      Session data is denormalised and tightly coupled to the Excel layout —
      importing it reliably would require significant reverse-engineering work
      and is prone to data quality issues.  Start fresh sessions from the web app.
      Roster entries are likewise skipped because they depend on sessions.
"""

import os
import sys
import re
from datetime import date
from pathlib import Path

import openpyxl
from supabase import create_client, Client

EXCEL_PATH = Path(__file__).resolve().parent.parent / "Badminton.xlsx"

# Valid skill_level values as defined in the schema
VALID_SKILL_LEVELS = {"HB", "LI", "MB"}

# Column index of each owner in the Shuttle Purchase header row.
# Owner columns come in pairs: (count, description). We only need the count column.
# These are parsed dynamically from the header row rather than hardcoded.


def get_client() -> Client:
    url = os.environ["SUPABASE_URL"]
    key = os.environ["SUPABASE_SERVICE_ROLE_KEY"]
    return create_client(url, key)


def _warn_missing_sheet(name: str) -> None:
    print(f"  WARNING: sheet '{name}' not found — skipping.")


def _normalise_skill_level(raw: str | None) -> str | None:
    """Return the raw value if it is a valid schema skill level, else None."""
    if raw is None:
        return None
    stripped = str(raw).strip()
    if stripped in VALID_SKILL_LEVELS:
        return stripped
    return None


def _extract_brand(shuttle_type: str) -> str:
    """Return the brand portion before any parenthesised date/qualifier."""
    return re.split(r"\s*\(", shuttle_type.strip())[0].strip()


def import_venues(ws, client: Client) -> tuple[int, int]:
    """
    Import venues from the 'Court List' sheet.

    Expected layout (row numbers are 1-based):
      Row 1: title row  ("Court List", ...)
      Row 2: blank
      Row 3: headers    ("Location", "Court Cost...", "Amount Paid by Pubs...")
      Row 4+: data
    """
    # Locate the header row (contains "Location")
    header_row_idx = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        if row[0] == "Location":
            header_row_idx = i
            break
    if header_row_idx is None:
        print("  WARNING: could not locate header row in 'Court List' — skipping.")
        return 0, 0

    # Read all columns from header row
    headers = [str(h).strip() if h is not None else "" for h in
               next(ws.iter_rows(min_row=header_row_idx, max_row=header_row_idx, values_only=True))]

    # Identify column indices
    col_name = headers.index("Location")
    cost_col = next((i for i, h in enumerate(headers) if "Court Cost" in h), None)
    pub_fee_col = next((i for i, h in enumerate(headers) if "Amount Paid by Pubs" in h), None)

    if cost_col is None or pub_fee_col is None:
        print("  WARNING: expected cost columns not found in 'Court List' — skipping.")
        return 0, 0

    # Fetch existing venue names to avoid duplicates
    existing = {r["name"] for r in client.table("venues").select("name").execute().data}

    inserted = 0
    skipped = 0
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        name = row[col_name]
        if not name or not isinstance(name, str):
            continue
        name = name.strip()
        if not name:
            continue

        court_cost = row[cost_col]
        pub_fee = row[pub_fee_col]

        if court_cost is None or pub_fee is None:
            skipped += 1
            continue

        if name in existing:
            skipped += 1
            continue

        client.table("venues").insert({
            "name": name,
            "court_cost_per_hour": float(court_cost),
            "default_pub_fee": float(pub_fee),
        }).execute()
        existing.add(name)
        inserted += 1

    return inserted, skipped


def import_members(ws, client: Client) -> tuple[int, int]:
    """
    Import internal team members from the 'Member List' sheet.

    The sheet is a session-attendance log (dates × member columns).
    We extract all unique non-empty names that appear in the member columns
    (columns 3–10 based on the header row) and import them as internal players.

    All imported players get is_internal=True, is_admin=False, skill_level=NULL
    (admins can update skill levels in the app).
    """
    # Locate header row (first row where col[0] == "Date")
    header_row_idx = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        if row[0] == "Date":
            header_row_idx = i
            break
    if header_row_idx is None:
        print("  WARNING: could not locate header row in 'Member List' — skipping.")
        return 0, 0

    # Member name columns: everything from col index 2 onward that has a header
    headers = list(next(ws.iter_rows(
        min_row=header_row_idx, max_row=header_row_idx, values_only=True
    )))
    member_col_indices = [
        i for i, h in enumerate(headers)
        if h is not None and isinstance(h, str) and h.strip().startswith("Member")
    ]

    # Collect unique names, excluding placeholder column-header literals
    placeholder_re = re.compile(r"^Member\s+\d+$", re.IGNORECASE)
    unique_names: set[str] = set()
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if row[0] is None:
            continue
        for idx in member_col_indices:
            if idx < len(row) and row[idx] and isinstance(row[idx], str):
                name = row[idx].strip()
                if name and not placeholder_re.match(name):
                    unique_names.add(name)

    # Fetch existing player names to avoid duplicates
    existing = {r["name"] for r in client.table("players").select("name").execute().data}

    inserted = 0
    skipped = 0
    for name in sorted(unique_names):
        if name in existing:
            skipped += 1
            continue
        client.table("players").insert({
            "name": name,
            "is_internal": True,
            "is_admin": False,
        }).execute()
        existing.add(name)
        inserted += 1

    return inserted, skipped


def import_pub_players(ws, client: Client) -> tuple[int, int]:
    """
    Import public (external) players from the 'Pub List-Selection' sheet.

    Expected layout:
      Row 1: title ("Pub List - Selection")
      Row 2: headers ("Pubs", "Skill Level", "Toxicity Level ...", "Description")
      Row 3+: data

    Skill level is imported only when the raw value matches a valid schema value
    (HB, LI, MB); other values are discarded.
    The toxicity level column is stored as the player's notes field if present.
    """
    # Locate header row (first row where col[0] in {"Pubs", "Pub List - Selection"})
    header_row_idx = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), start=1):
        if row[0] == "Pubs":
            header_row_idx = i
            break
    if header_row_idx is None:
        print("  WARNING: could not locate header row in 'Pub List-Selection' — skipping.")
        return 0, 0

    headers = [str(h).strip() if h is not None else "" for h in
               next(ws.iter_rows(min_row=header_row_idx, max_row=header_row_idx, values_only=True))]

    col_name = headers.index("Pubs")
    skill_col = next((i for i, h in enumerate(headers) if h == "Skill Level"), None)
    notes_col = next((i for i, h in enumerate(headers) if "Toxicity" in h), None)

    existing = {r["name"] for r in client.table("players").select("name").execute().data}

    inserted = 0
    skipped = 0
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        name = row[col_name]
        if not name or not isinstance(name, str):
            continue
        name = name.strip()
        if not name:
            continue

        if name in existing:
            skipped += 1
            continue

        skill = _normalise_skill_level(row[skill_col] if skill_col is not None else None)
        notes = str(row[notes_col]).strip() if (
            notes_col is not None and row[notes_col] is not None
        ) else None

        record = {
            "name": name,
            "is_internal": False,
            "is_admin": False,
        }
        if skill is not None:
            record["skill_level"] = skill
        if notes:
            record["notes"] = notes

        client.table("players").insert(record).execute()
        existing.add(name)
        inserted += 1

    return inserted, skipped


def import_shuttle_batches(ws, client: Client) -> tuple[int, int]:
    """
    Import shuttle purchase history from the 'Shuttle Purchase' sheet.

    Layout:
      Row 1: title
      Row 2: headers — "Shuttle Type", "Shuttle Cost ($ per Tube)",
                        "Shuttle Cost ($ per Shuttle)", "Purchased Date",
                        then repeating pairs: "<Owner> (No. of Shuttles)", None
      Row 3+: data

    One shuttle_batch row is created per (shuttle_type × owner) combination
    where the owner's count is a valid number (non-null, not '-').

    All imported batches are marked is_active=False (historical stock).
    Admins can activate a batch and adjust remaining_count in the web app.

    shuttles_per_tube is derived from cost_per_tube / cost_per_shuttle,
    rounded to the nearest integer (all rows in the file use 12 shuttles/tube).
    """
    # Locate header row
    header_row_idx = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), start=1):
        if row[0] == "Shuttle Type":
            header_row_idx = i
            break
    if header_row_idx is None:
        print("  WARNING: could not locate header row in 'Shuttle Purchase' — skipping.")
        return 0, 0

    headers = list(next(ws.iter_rows(
        min_row=header_row_idx, max_row=header_row_idx, values_only=True
    )))

    # Fixed column indices
    col_type = 0          # "Shuttle Type"
    col_cost_tube = 1     # "Shuttle Cost ($ per Tube)"
    col_cost_shuttle = 2  # "Shuttle Cost ($ per Shuttle)"
    col_date = 3          # "Purchased Date"

    # Parse owner columns: each owner occupies two adjacent columns
    # (count column has header "X (No. of Shuttles)", description column is None).
    owner_cols: list[tuple[str, int]] = []  # [(owner_label, col_index), ...]
    for i, h in enumerate(headers):
        if h is not None and isinstance(h, str) and "(No. of Shuttles)" in h:
            owner_label = h.replace("(No. of Shuttles)", "").strip()
            owner_cols.append((owner_label, i))

    # Build set of existing (batch_name, owner_label) pairs to detect duplicates
    existing_rows = client.table("shuttle_batches").select("batch_name,owner_label").execute().data
    existing: set[tuple[str, str]] = {(r["batch_name"], r["owner_label"]) for r in existing_rows}

    inserted = 0
    skipped = 0

    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        shuttle_type = row[col_type]
        if not shuttle_type or not isinstance(shuttle_type, str):
            continue
        shuttle_type = shuttle_type.strip()

        cost_per_tube = row[col_cost_tube]
        cost_per_shuttle = row[col_cost_shuttle]
        purchased_date = row[col_date]

        if cost_per_tube is None or cost_per_shuttle is None:
            continue

        # Derive shuttles_per_tube
        try:
            shuttles_per_tube = round(float(cost_per_tube) / float(cost_per_shuttle))
        except (ZeroDivisionError, TypeError, ValueError):
            shuttles_per_tube = 12  # sensible default for the known data

        # Convert purchased_date to ISO string if it's a datetime
        purchased_at = None
        if purchased_date is not None:
            if hasattr(purchased_date, "date"):
                purchased_at = purchased_date.date().isoformat()
            elif isinstance(purchased_date, date):
                purchased_at = purchased_date.isoformat()

        brand = _extract_brand(shuttle_type)
        batch_name = shuttle_type  # full name including date qualifier

        for owner_label, col_idx in owner_cols:
            if col_idx >= len(row):
                continue
            raw_count = row[col_idx]
            # Skip missing or dash placeholder values
            if raw_count is None or str(raw_count).strip() in {"-", ""}:
                continue
            try:
                count = int(float(raw_count))
            except (TypeError, ValueError):
                continue

            key = (batch_name, owner_label)
            if key in existing:
                skipped += 1
                continue

            client.table("shuttle_batches").insert({
                "batch_name": batch_name,
                "brand": brand,
                "owner_label": owner_label,
                "cost_per_tube": float(cost_per_tube),
                "shuttles_per_tube": shuttles_per_tube,
                "remaining_count": count,
                "is_active": False,
                **({"purchased_at": purchased_at} if purchased_at else {}),
            }).execute()
            existing.add(key)
            inserted += 1

    return inserted, skipped


def main() -> None:
    if not EXCEL_PATH.exists():
        print(f"ERROR: Excel file not found at {EXCEL_PATH}", file=sys.stderr)
        sys.exit(1)

    client = get_client()

    print(f"Loading workbook from {EXCEL_PATH} …")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    print(f"Sheets found: {wb.sheetnames}\n")

    # ── Venues ────────────────────────────────────────────────────────────────
    print("Importing venues (Court List) …")
    if "Court List" in wb.sheetnames:
        ins, skp = import_venues(wb["Court List"], client)
        print(f"  {ins} inserted, {skp} skipped")
    else:
        _warn_missing_sheet("Court List")

    # ── Internal members ──────────────────────────────────────────────────────
    print("\nImporting internal members (Member List) …")
    if "Member List" in wb.sheetnames:
        ins, skp = import_members(wb["Member List"], client)
        print(f"  {ins} inserted, {skp} skipped")
    else:
        _warn_missing_sheet("Member List")

    # ── Public players ────────────────────────────────────────────────────────
    print("\nImporting public players (Pub List-Selection) …")
    if "Pub List-Selection" in wb.sheetnames:
        ins, skp = import_pub_players(wb["Pub List-Selection"], client)
        print(f"  {ins} inserted, {skp} skipped")
    else:
        _warn_missing_sheet("Pub List-Selection")

    # ── Shuttle batches ───────────────────────────────────────────────────────
    print("\nImporting shuttle purchase history (Shuttle Purchase) …")
    if "Shuttle Purchase" in wb.sheetnames:
        ins, skp = import_shuttle_batches(wb["Shuttle Purchase"], client)
        print(f"  {ins} inserted, {skp} skipped")
    else:
        _warn_missing_sheet("Shuttle Purchase")

    print("\nImport complete.")


if __name__ == "__main__":
    main()
